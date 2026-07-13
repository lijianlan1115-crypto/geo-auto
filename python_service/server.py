import base64
import json
import os
import re
import shutil
import sqlite3
import threading
import time
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from config import (
    ANSWER_POLL_INTERVAL,
    ANSWER_KEYWORD_STABLE_SECONDS,
    ANSWER_STABLE_SECONDS,
    ANSWER_TIMEOUT_SECONDS,
    CONCURRENCY,
    DB_PATH,
    HOST,
    ID_HEADERS,
    INPUT_EXCEL,
    KEYWORD,
    KEYWORD_HEADERS,
    KEYWORDS,
    MAX_FOLLOWUPS,
    OUTPUT_DIR,
    PLATFORM_COLUMN_ALIASES,
    PLATFORMS,
    PORT,
    QUESTION_HEADERS,
    RESULT_EXCEL,
    TEMP_ANSWERS_EXCEL,
    SCREENSHOT_DIR,
)

from ocr_checker import check_keyword
from image_marker import mark_image
from ai_judge import ai_judge, configure_ai_judge, generate_followup, get_ai_judge_config

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
except ImportError as exc:
    raise SystemExit("缺少 openpyxl，请先安装：pip install openpyxl") from exc


lock = threading.Lock()


class ReusableThreadingHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_dirs():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    for platform in PLATFORMS:
        (SCREENSHOT_DIR / platform).mkdir(parents=True, exist_ok=True)


def save_workbook_atomic(workbook, target_path):
    target = Path(target_path)
    temp = target.with_name(f".{target.stem}.{os.getpid()}.tmp.xlsx")
    try:
        workbook.save(temp)
        os.replace(temp, target)
    finally:
        if temp.exists():
            temp.unlink()


def connect_db():
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    ensure_dirs()
    with connect_db() as conn:
        conn.execute("pragma journal_mode=WAL")
        conn.execute("pragma synchronous=FULL")
        conn.execute(
            """
            create table if not exists tasks (
                task_id text primary key,
                row_number integer not null,
                row_id text,
                question text not null,
                platform text not null,
                status text not null default 'pending',
                matched integer,
                followup_count integer default 0,
                screenshot_path text,
                answer_text text,
                error text,
                created_at text not null,
                updated_at text not null
            )
            """
        )
        conn.execute("create index if not exists idx_tasks_status on tasks(status)")
        existing = {row[1] for row in conn.execute("pragma table_info(tasks)").fetchall()}
        if "answer_debug" not in existing:
            conn.execute("alter table tasks add column answer_debug text")
        if "run_debug" not in existing:
            conn.execute("alter table tasks add column run_debug text")


def read_headers(sheet):
    headers = {}
    for cell in sheet[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    return headers


def find_column(headers, names):
    for name in names:
        if name in headers:
            return headers[name]
    return None


def find_platform_column(headers, platform):
    names = [PLATFORMS[platform]["column"]]
    names.extend(PLATFORM_COLUMN_ALIASES.get(platform, []))
    return find_column(headers, names)


def split_keywords(value):
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[\n,，、;；|/]+|\s+or\s+|\s+OR\s+", text)
    keywords = [part.strip() for part in parts if part and part.strip()]
    return keywords


def safe_platform_key(value, index=0):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9_\-]+", "_", text)
    text = text.strip("_")
    return text or f"custom_{index + 1}"


def configure_platforms(platforms):
    if not isinstance(platforms, list) or not platforms:
        return {"ok": False, "error": "平台列表为空"}

    new_platforms = {}
    used_keys = set()
    for index, item in enumerate(platforms):
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or "").strip()
        if not url:
            continue
        base_key = safe_platform_key(item.get("key") or item.get("name"), index)
        key = base_key
        duplicate_index = 2
        while key in used_keys:
            key = f"{base_key}_{duplicate_index}"
            duplicate_index += 1
        used_keys.add(key)
        name = str(item.get("name") or key).strip()
        existing = PLATFORMS.get(key)
        new_platforms[key] = {
            "name": name,
            "column": existing.get("column") if existing else name,
            "url": url,
        }

    if not new_platforms:
        return {"ok": False, "error": "没有有效的平台 URL"}

    PLATFORMS.clear()
    PLATFORMS.update(new_platforms)
    ensure_dirs()
    seed_tasks()
    with connect_db() as conn:
        placeholders = ",".join("?" for _ in new_platforms)
        conn.execute(
            f"delete from tasks where platform not in ({placeholders})",
            tuple(new_platforms.keys()),
        )
    return {"ok": True, "platforms": runtime_platforms(), "stats": stats()}


def runtime_platforms():
    return [
        {"key": key, "name": item["name"], "url": item["url"]}
        for key, item in PLATFORMS.items()
    ]


def clear_previous_outputs(ws, headers):
    output_headers = {
        "豆包", "千问", "DeepSeek", "deepseek", "元宝", "文心一言",
        "豆包_状态", "千问_状态", "DeepSeek_状态", "deepseek_状态", "元宝_状态", "文心一言_状态",
        "豆包_追问次数", "千问_追问次数", "DeepSeek_追问次数", "deepseek_追问次数", "元宝_追问次数", "文心一言_追问次数",
    }
    for item in PLATFORMS.values():
        output_headers.add(item["column"])
        output_headers.add(f"{item['column']}_状态")
        output_headers.add(f"{item['column']}_追问次数")

    clear_cols = []
    for header, col in headers.items():
      text = str(header or "").strip()
      if text in output_headers or text.endswith("_状态") or text.endswith("_追问次数"):
          clear_cols.append(col)

    for row in range(2, ws.max_row + 1):
        for col in clear_cols:
            ws.cell(row=row, column=col, value=None)
    if hasattr(ws, "_images"):
        ws._images = []


def prepare_workbook(clear_outputs=False):
    if not INPUT_EXCEL.exists():
        raise FileNotFoundError(f"找不到输入 Excel：{INPUT_EXCEL}")

    if not RESULT_EXCEL.exists():
        shutil.copy2(INPUT_EXCEL, RESULT_EXCEL)

    wb = load_workbook(RESULT_EXCEL)
    ws = wb.active
    headers = read_headers(ws)

    if clear_outputs:
        clear_previous_outputs(ws, headers)

    question_col = find_column(headers, QUESTION_HEADERS)
    if question_col is None:
        detected = ", ".join(headers.keys()) or "空表头"
        raise RuntimeError(
            "Excel 第一行找不到问题列。"
            f"支持列名：{QUESTION_HEADERS}。"
            f"当前检测到的表头：{detected}。"
            "请使用包含“id”和“问题”列的原始问题表作为 input.xlsx。"
        )

    id_col = find_column(headers, ID_HEADERS)
    keyword_col = find_column(headers, KEYWORD_HEADERS)

    max_col = ws.max_column
    for platform, item in PLATFORMS.items():
        column_name = item["column"]
        existing_col = find_platform_column(headers, platform)
        if existing_col is None:
            max_col += 1
            ws.cell(row=1, column=max_col, value=column_name)
            headers[column_name] = max_col
        elif column_name not in headers:
            headers[column_name] = existing_col

        status_name = f"{column_name}_状态"
        if status_name not in headers:
            max_col += 1
            ws.cell(row=1, column=max_col, value=status_name)
            headers[status_name] = max_col

        followup_name = f"{column_name}_追问次数"
        if followup_name not in headers:
            max_col += 1
            ws.cell(row=1, column=max_col, value=followup_name)
            headers[followup_name] = max_col

    save_workbook_atomic(wb, RESULT_EXCEL)
    return question_col, id_col, keyword_col


def seed_tasks(clear_outputs=False):
    try:
        question_col, id_col, keyword_col = prepare_workbook(clear_outputs=clear_outputs)
    except Exception as exc:
        print(f"任务初始化跳过：{exc}")
        print("服务会继续启动，方便先测试 Chrome 插件连接；换成包含“问题”列的 Excel 后再重启即可生成任务。")
        return 0

    wb = load_workbook(RESULT_EXCEL, read_only=True)
    ws = wb.active

    created = 0
    with connect_db() as conn:
        for row_number in range(2, ws.max_row + 1):
            question = ws.cell(row=row_number, column=question_col).value
            if question is None or not str(question).strip():
                continue

            row_id = ws.cell(row=row_number, column=id_col).value if id_col else row_number
            row_keywords = split_keywords(ws.cell(row=row_number, column=keyword_col).value if keyword_col else None)
            for platform in PLATFORMS:
                task_id = f"{row_number}:{platform}"
                exists = conn.execute(
                    "select 1 from tasks where task_id = ?", (task_id,)
                ).fetchone()
                if exists:
                    continue

                conn.execute(
                    """
                    insert into tasks (
                        task_id, row_number, row_id, question, platform, status,
                        answer_text, created_at, updated_at
                    )
                    values (?, ?, ?, ?, ?, 'pending', ?, ?, ?)
                    """,
                    (
                        task_id,
                        row_number,
                        str(row_id),
                        str(question).strip(),
                        platform,
                        json.dumps({"keywords": row_keywords}, ensure_ascii=False),
                        now(),
                        now(),
                    ),
                )
                created += 1
    return created


def get_next_task():
    with lock, connect_db() as conn:
        task = conn.execute(
            """
            select * from tasks
            where status = 'pending'
            order by row_number asc, platform asc
            limit 1
            """
        ).fetchone()

        if not task:
            return None

        conn.execute(
            "update tasks set status = 'running', updated_at = ? where task_id = ?",
            (now(), task["task_id"]),
        )

    platform = PLATFORMS[task["platform"]]
    keywords = []
    try:
        meta = json.loads(task["answer_text"] or "{}")
        keywords = meta.get("keywords") or []
    except Exception:
        pass
    return {
        "task_id": task["task_id"],
        "row_number": task["row_number"],
        "row_id": task["row_id"],
        "question": task["question"],
        "keyword": keywords[0] if keywords else "",
        "keywords": keywords,
        "platform": task["platform"],
        "platform_name": platform["name"],
        "platform_url": platform["url"],
        "max_followups": MAX_FOLLOWUPS,
        "answer_poll_interval": ANSWER_POLL_INTERVAL,
        "answer_stable_seconds": ANSWER_STABLE_SECONDS,
        "answer_keyword_stable_seconds": ANSWER_KEYWORD_STABLE_SECONDS,
        "answer_timeout_seconds": ANSWER_TIMEOUT_SECONDS,
    }


def image_from_data_url(data_url, platform, task_id, matched):
    if not data_url:
        return None

    if "," in data_url:
        _, raw = data_url.split(",", 1)
    else:
        raw = data_url

    suffix = "hit" if matched else "miss"
    safe_task_id = task_id.replace(":", "_")
    path = SCREENSHOT_DIR / platform / f"{safe_task_id}_{suffix}.png"
    path.parent.mkdir(parents=True, exist_ok=True)

    # 覆盖旧截图，确保每次重新执行都得到最新图片
    if path.exists():
        path.unlink()
    path.write_bytes(base64.b64decode(raw))
    return path



def process_screenshot_with_ocr(image_path, keywords, region_ratio=None):
    """
    截图二次检测：
    OCR识别关键词并画框
    """

    if not image_path:
        return {
            "matched": False
        }

    try:
        result = check_keyword(
            image_path,
            keywords,
            region_ratio=region_ratio,
        )

        if result.get("matched"):

            bbox = result.get("bbox")

            if bbox:
                mark_image(
                    image_path,
                    bbox
                )

        return result

    except Exception as e:
        print("OCR处理失败:", e)

        return {
            "matched": False,
            "error": str(e)
        }


def save_test_screenshot(payload):
    data_url = payload.get("screenshot_data_url")
    if not data_url:
        raise ValueError("缺少 screenshot_data_url")

    if "," in data_url:
        _, raw = data_url.split(",", 1)
    else:
        raw = data_url

    test_dir = SCREENSHOT_DIR / "test"
    test_dir.mkdir(parents=True, exist_ok=True)
    keywords = payload.get("keywords") or split_keywords(payload.get("keyword") or KEYWORD)
    safe_keyword = "_".join("".join(ch for ch in item if ch.isalnum() or ch in ("-", "_")) for item in keywords)
    safe_keyword = safe_keyword[:80] or "keyword"
    path = test_dir / f"test_{safe_keyword}_{int(time.time())}.png"
    path.write_bytes(base64.b64decode(raw))
    return {"ok": True, "screenshot_path": str(path)}


def get_test_keywords():
    if not INPUT_EXCEL.exists():
        return {
            "ok": True,
            "source": "missing",
            "keywords": [],
            "message": f"找不到输入 Excel：{INPUT_EXCEL}",
        }

    wb = load_workbook(INPUT_EXCEL, read_only=True)
    ws = wb.active
    headers = read_headers(ws)
    keyword_col = find_column(headers, KEYWORD_HEADERS)
    question_col = find_column(headers, QUESTION_HEADERS)

    if keyword_col is None:
        return {
            "ok": True,
            "source": "plugin",
            "keywords": [],
            "message": f"Excel 没有关键词列，支持列名：{KEYWORD_HEADERS}",
        }

    for row_number in range(2, ws.max_row + 1):
        keywords = split_keywords(ws.cell(row=row_number, column=keyword_col).value)
        question = ws.cell(row=row_number, column=question_col).value if question_col else None
        if keywords:
            return {
                "ok": True,
                "source": "excel",
                "row_number": row_number,
                "question": question,
                "keywords": keywords,
            }

    return {
        "ok": True,
        "source": "plugin",
        "keywords": [],
        "message": "Excel 关键词列为空，请使用插件面板中填写的目标关键词",
    }


def col_letter(ws, column_index):
    return ws.cell(row=1, column=column_index).column_letter


def image_anchor_cell(image):
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        return anchor

    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None

    row = int(marker.row) + 1
    col = int(marker.col) + 1
    return f"{col_letter_from_index(col)}{row}"


def col_letter_from_index(column_index):
    letters = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def remove_images_at_cell(ws, cell_ref):
    kept = []
    for image in getattr(ws, "_images", []):
        if image_anchor_cell(image) != cell_ref:
            kept.append(image)
    ws._images = kept


def anchor_image_inside_cell(img, row_number, column_index):
    img.anchor = TwoCellAnchor(
        editAs="twoCell",
        _from=AnchorMarker(col=column_index - 1, row=row_number - 1),
        to=AnchorMarker(col=column_index, row=row_number),
    )
    # Excel stores images as drawing objects. This keeps them visually bound to
    # the cell area instead of behaving like an arbitrary floating picture.
    img.object_position = 1


def migrate_images_inside_cells(ws):
    changed = False
    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        if isinstance(anchor, TwoCellAnchor) and getattr(anchor, "editAs", None) == "twoCell":
            continue
        anchor_image_inside_cell(img, int(marker.row) + 1, int(marker.col) + 1)
        changed = True
    return changed


def compact_text(value, limit=None):
    text = str(value or "").replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit] if limit else text


def write_temp_answer_sheet(payload):
    headers = [
        "更新时间", "task_id", "行号", "平台", "状态", "是否命中", "命中关键词",
        "追问次数", "回答长度", "题目", "目标关键词", "回答片段", "完整回答", "每轮调试"
    ]
    TEMP_ANSWERS_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    if TEMP_ANSWERS_EXCEL.exists():
        wb = load_workbook(TEMP_ANSWERS_EXCEL)
        ws = wb.active
        existing_headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        if existing_headers[:len(headers)] != headers:
            ws.insert_rows(1)
            for col, name in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AI返回内容"
        for col, name in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=name)

    task_id = str(payload.get("task_id") or "")
    target_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value or "") == task_id:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1

    answer_text = str(payload.get("answer_text") or "")
    run_debug = payload.get("run_debug") or []
    matched_keywords = payload.get("matched_keywords") or []
    keywords = payload.get("keywords") or payload.get("keyword") or ""
    if isinstance(keywords, list):
        keywords_text = "，".join(str(item) for item in keywords)
    else:
        keywords_text = str(keywords or "")

    values = [
        now(),
        task_id,
        payload.get("row_number"),
        payload.get("platform"),
        payload.get("status", "done"),
        "是" if payload.get("matched") else "否",
        "，".join(str(item) for item in matched_keywords),
        int(payload.get("followup_count", 0) or 0),
        len(answer_text),
        compact_text(payload.get("question"), 500),
        keywords_text,
        compact_text(answer_text, 1000),
        answer_text[:30000],
        json.dumps(run_debug, ensure_ascii=False)[:30000],
    ]
    for col, value in enumerate(values, 1):
        ws.cell(row=target_row, column=col, value=value)

    widths = {1: 20, 2: 18, 4: 14, 6: 10, 7: 24, 10: 42, 11: 26, 12: 70, 13: 90, 14: 90}
    for col, width in widths.items():
        ws.column_dimensions[col_letter(ws, col)].width = max(ws.column_dimensions[col_letter(ws, col)].width or 0, width)
    wb.save(TEMP_ANSWERS_EXCEL)


def write_result_to_excel(result, screenshot_path):
    wb = load_workbook(RESULT_EXCEL)
    ws = wb.active
    headers = read_headers(ws)
    migrate_images_inside_cells(ws)

    platform = result["platform"]
    platform_config = PLATFORMS[platform]
    image_col = headers[platform_config["column"]]
    status_col = headers[f"{platform_config['column']}_状态"]
    followup_col = headers[f"{platform_config['column']}_追问次数"]
    row_number = int(result["row_number"])

    matched = bool(result.get("matched"))
    matched_keywords = result.get("matched_keywords") or []
    error_text = str(result.get("error") or "").strip()
    if error_text:
        status_text = f"失败：{compact_text(error_text, 60)}"
    elif matched and matched_keywords:
        status_text = f"命中：{'，'.join(matched_keywords)}"
    else:
        status_text = "命中" if matched else "未命中"

    ws.cell(row=row_number, column=status_col, value=status_text)
    ws.cell(row=row_number, column=followup_col, value=int(result.get("followup_count", 0)))

    image_cell = f"{col_letter(ws, image_col)}{row_number}"
    remove_images_at_cell(ws, image_cell)

    if screenshot_path and Path(screenshot_path).exists():
        img = ExcelImage(str(screenshot_path))
        original_width = max(1, int(getattr(img, "width", 1) or 1))
        original_height = max(1, int(getattr(img, "height", 1) or 1))
        max_width = 520
        max_height = 330
        scale = min(max_width / original_width, max_height / original_height, 1.0)
        if original_width > max_width or original_height > max_height:
            img.width = int(original_width * scale)
            img.height = int(original_height * scale)

        ws.row_dimensions[row_number].height = max(ws.row_dimensions[row_number].height or 0, int(img.height * 0.75) + 12)
        ws.column_dimensions[col_letter(ws, image_col)].width = max(
            ws.column_dimensions[col_letter(ws, image_col)].width or 0, min(76, max(38, img.width / 7))
        )
        anchor_image_inside_cell(img, row_number, image_col)
        ws.add_image(img)

    save_workbook_atomic(wb, RESULT_EXCEL)


def submit_result(payload):
    required = ["task_id", "row_number", "platform"]
    missing = [key for key in required if key not in payload]
    if missing:
        raise ValueError(f"缺少字段：{missing}")

    # V2.0：优先使用浏览器 DOM 定位结果
    dom_location = payload.get("dom_location") or {}
    dom_matched = bool(dom_location.get("matched"))
    dom_keywords = dom_location.get("matched_keywords") or []

    content_matched = bool(payload.get("matched"))
    content_keywords = payload.get("matched_keywords") or []

    # 如果 DOM 定位成功，优先采用 DOM 结果
    if dom_matched and dom_keywords:
        matched = True
        matched_keywords = dom_keywords
    else:
        matched = content_matched
        matched_keywords = content_keywords

    screenshot_path = image_from_data_url(
        payload.get("screenshot_data_url"),
        payload["platform"],
        payload["task_id"],
        matched,
    )

    if screenshot_path and payload.get("platform") == "qianwen" and matched:
        ocr_keywords = payload.get("keywords") or matched_keywords or content_keywords
        ocr_result = process_screenshot_with_ocr(
            screenshot_path,
            ocr_keywords,
            # 千问截图只在正文区域 OCR，避开左侧栏、顶部问题气泡和底部输入框。
            region_ratio=(0.18, 0.14, 0.96, 0.82),
        )
        payload["ocr_debug"] = ocr_result
        if ocr_result.get("matched") and ocr_result.get("keyword"):
            matched = True
            matched_keywords = list(dict.fromkeys([*matched_keywords, ocr_result.get("keyword")]))

    payload["matched"] = matched
    payload["matched_keywords"] = matched_keywords
    task_status = "failed" if payload.get("error") else "done"

    with lock:
        if screenshot_path:
            write_result_to_excel(payload, screenshot_path)
            # 图片已嵌入 Excel，删除本地截图文件避免堆积
            try:
                Path(screenshot_path).unlink()
            except Exception:
                pass

        with connect_db() as conn:
            conn.execute(
                """
                update tasks
                set status = ?,
                    matched = ?,
                    followup_count = ?,
                    screenshot_path = ?,
                    answer_text = ?,
                    answer_debug = ?,
                    run_debug = ?,
                    error = ?,
                    updated_at = ?
                where task_id = ?
                """,
                (
                    task_status,
                    1 if matched else 0,
                    int(payload.get("followup_count", 0)),
                    str(screenshot_path) if screenshot_path else None,
                    payload.get("answer_text", "")[:20000],
                    json.dumps(payload.get("answer_debug") or {}, ensure_ascii=False)[:20000],
                    json.dumps(payload.get("run_debug") or [], ensure_ascii=False)[:40000],
                    payload.get("error"),
                    now(),
                    payload["task_id"],
                ),
            )

    return {"ok": True, "screenshot_path": str(screenshot_path) if screenshot_path else None}


def judge_answer(payload):
    answer_text = payload.get("answer_text") or ""
    keywords = payload.get("keywords") or split_keywords(payload.get("keyword") or KEYWORD)
    if not isinstance(keywords, list):
        keywords = split_keywords(keywords)
    result = ai_judge(
        answer_text=answer_text,
        keywords=keywords,
        question=payload.get("question") or "",
        platform=payload.get("platform") or "",
    )
    result["answer_length"] = len(str(answer_text or ""))
    return result


def generate_followup_prompt(payload):
    keywords = payload.get("keywords") or split_keywords(payload.get("keyword") or KEYWORD)
    if not isinstance(keywords, list):
        keywords = split_keywords(keywords)
    return generate_followup(
        question=payload.get("question") or "",
        answer_text=payload.get("answer_text") or "",
        keywords=keywords,
        followup_count=payload.get("followup_count") or 0,
        platform=payload.get("platform") or "",
    )


def mark_failed(payload):
    task_id = payload.get("task_id")
    if not task_id:
        raise ValueError("缺少 task_id")

    failed_payload = dict(payload)
    failed_payload["status"] = "failed"
    failed_payload["matched"] = False
    write_temp_answer_sheet(failed_payload)

    with connect_db() as conn:
        conn.execute(
            """
            update tasks
            set status = 'failed', error = ?, answer_debug = ?, run_debug = ?, updated_at = ?
            where task_id = ?
            """,
            (
                payload.get("error", "unknown error"),
                json.dumps(payload.get("answer_debug") or {}, ensure_ascii=False)[:20000],
                json.dumps(payload.get("run_debug") or [], ensure_ascii=False)[:40000],
                now(),
                task_id,
            ),
        )
    return {"ok": True}


def recent_debug(limit=20):
    with connect_db() as conn:
        rows = conn.execute(
            """
            select task_id, row_number, platform, status, matched, followup_count,
                   substr(answer_text, 1, 600) as answer_preview, error,
                   answer_debug, run_debug, updated_at
            from tasks
            order by updated_at desc
            limit ?
            """,
            (int(limit),),
        ).fetchall()
    items = []
    for row in rows:
        item = dict(row)
        for key in ("answer_debug", "run_debug"):
            try:
                item[key] = json.loads(item.get(key) or "{}")
            except Exception:
                pass
        items.append(item)
    return {"ok": True, "items": items}


def stats():
    with connect_db() as conn:
        rows = conn.execute(
            "select status, count(*) as count from tasks group by status order by status"
        ).fetchall()
        return {row["status"]: row["count"] for row in rows}


def reset_failed_tasks():
    with connect_db() as conn:
        cursor = conn.execute(
            """
            update tasks
            set status = 'pending',
                matched = null,
                followup_count = 0,
                screenshot_path = null,
                error = null,
                updated_at = ?
            where status in ('failed', 'running')
            """,
            (now(),),
        )
        reset_count = cursor.rowcount
    return {"ok": True, "reset_count": reset_count, "stats": stats()}


def reset_all_tasks():
    with connect_db() as conn:
        conn.execute("delete from tasks")
    if RESULT_EXCEL.exists():
        RESULT_EXCEL.unlink()
    created = seed_tasks(clear_outputs=True)
    return {"ok": True, "reset_count": created, "created": created, "stats": stats()}


class Handler(BaseHTTPRequestHandler):
    def _send(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _json_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def _send_html(self, html_text):
        body = html_text.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self._send(200, {"ok": True})

    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/health":
            self._send(200, {"ok": True, "keyword": KEYWORD, "keywords": KEYWORDS, "stats": stats()})
        elif path == "/config":
            self._send(
                200,
                {
                    "keyword": KEYWORD,
                    "keywords": KEYWORDS,
                    "max_followups": MAX_FOLLOWUPS,
                    "concurrency": CONCURRENCY,
                    "platforms": runtime_platforms(),
                    "ai_judge": get_ai_judge_config(),
                    "answer_poll_interval": ANSWER_POLL_INTERVAL,
                    "answer_stable_seconds": ANSWER_STABLE_SECONDS,
                    "answer_keyword_stable_seconds": ANSWER_KEYWORD_STABLE_SECONDS,
                    "answer_timeout_seconds": ANSWER_TIMEOUT_SECONDS,
                },
            )
        elif path == "/next-task":
            task = get_next_task()
            self._send(200, {"ok": True, "task": task})
        elif path == "/test-keywords":
            self._send(200, get_test_keywords())
        elif path == "/ai-judge-config":
            self._send(200, get_ai_judge_config())
        elif path == "/debug/recent":
            query = urlparse(self.path).query
            limit = 20
            if query.startswith("limit="):
                try:
                    limit = int(query.split("=", 1)[1])
                except Exception:
                    pass
            self._send(200, recent_debug(limit))
        elif path == "/test-page":
            self._send_html(
                """
                <!doctype html>
                <html lang="zh-CN">
                  <head>
                    <meta charset="utf-8">
                    <title>GEO截图测试页</title>
                    <style>
                      body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; padding: 48px 80px; line-height: 1.8; }
                      main { max-width: 880px; margin: auto; }
                      h1 { font-size: 28px; }
                      p { font-size: 18px; }
                      .answer { border: 1px solid #dadce0; border-radius: 10px; padding: 24px; background: #fff; }
                    </style>
                  </head>
                  <body>
                    <main>
                      <h1>GEO截图测试页</h1>
                      <div class="answer">
                        <p>这是一个用于测试插件截图和关键词标注的页面。</p>
                        <p>示例回答：如果问题涉及贵州地区的商科、管理类、应用型本科院校，可以重点关注贵阳商学院，并结合专业方向、学费和就业城市进行比较。</p>
                        <p>点击左侧 GEO 小圆球，输入目标关键词后点击“测试截图”。</p>
                      </div>
                    </main>
                  </body>
                </html>
                """
            )
        elif path == "/mock-platform":
            self._send_html(
                """
                <!doctype html>
                <html lang="zh-CN">
                  <head>
                    <meta charset="utf-8">
                    <title>GEO模拟AI平台</title>
                    <style>
                      body {
                        margin: 0;
                        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
                        background: #f6f7f9;
                        color: #202124;
                      }
                      main { max-width: 920px; margin: 0 auto; padding: 40px 32px 120px; }
                      h1 { margin: 0 0 18px; font-size: 24px; }
                      .chat { display: grid; gap: 14px; }
                      .message {
                        border: 1px solid #dadce0;
                        border-radius: 10px;
                        padding: 16px 18px;
                        background: white;
                        line-height: 1.8;
                        font-size: 16px;
                      }
                      .user { background: #eef4ff; }
                      .composer {
                        position: fixed;
                        left: 50%;
                        bottom: 20px;
                        transform: translateX(-50%);
                        display: flex;
                        gap: 10px;
                        width: min(920px, calc(100vw - 64px));
                        padding: 12px;
                        border: 1px solid #dadce0;
                        border-radius: 12px;
                        background: white;
                        box-shadow: 0 12px 32px rgba(60, 64, 67, .18);
                      }
                      textarea { flex: 1; min-height: 48px; resize: vertical; border: 1px solid #dadce0; border-radius: 8px; padding: 10px; font-size: 15px; }
                      button { width: 88px; border: 0; border-radius: 8px; background: #1a73e8; color: white; font-size: 15px; cursor: pointer; }
                    </style>
                  </head>
                  <body>
                    <main>
                      <h1>GEO模拟AI平台</h1>
                      <section class="chat" id="chat">
                        <div class="message">这是本地模拟平台，用来测试插件自动提问、关键词标注、截图和 Excel 回写。</div>
                      </section>
                    </main>
                    <form class="composer" id="composer">
                      <textarea placeholder="请输入问题"></textarea>
                      <button type="submit">发送</button>
                    </form>
                    <script>
                      const chat = document.getElementById('chat');
                      const form = document.getElementById('composer');
                      const textarea = form.querySelector('textarea');
                      let count = 0;

                      function addMessage(text, cls) {
                        const div = document.createElement('div');
                        div.className = 'message ' + (cls || '');
                        div.textContent = text;
                        chat.appendChild(div);
                        div.scrollIntoView({ block: 'center' });
                      }

                      form.addEventListener('submit', (event) => {
                        event.preventDefault();
                        const question = textarea.value.trim();
                        if (!question) return;
                        textarea.value = '';
                        count += 1;
                        addMessage(question, 'user');
                        setTimeout(() => {
                          if (question.includes('第二') || question.includes('生活费')) {
                            addMessage('模拟回答：这个问题可以比较贵州商学院、贵阳学院等学校。这里故意命中第二个关键词贵州商学院，用来测试 OR 关系。');
                          } else if (count >= 6) {
                            addMessage('模拟回答：经过多轮追问后，仍然没有推荐目标学校，用来测试未命中截图。');
                          } else {
                            addMessage('模拟回答：结合贵州地区应用型本科和商科方向，可以重点关注贵阳商学院。这里包含目标关键词，插件应该滚动到这里并用红框标注。');
                          }
                        }, 800);
                      });
                    </script>
                  </body>
                </html>
                """
            )
        else:
            self._send(404, {"ok": False, "error": "not found"})

    def do_POST(self):
        path = urlparse(self.path).path
        try:
            payload = self._json_body()
            if path == "/submit-result":
                self._send(200, submit_result(payload))
            elif path == "/judge-answer":
                self._send(200, judge_answer(payload))
            elif path == "/generate-followup":
                self._send(200, generate_followup_prompt(payload))
            elif path == "/ai-judge-config":
                self._send(200, configure_ai_judge(payload))
            elif path == "/set-platforms":
                self._send(200, configure_platforms(payload.get("platforms") or []))
            elif path == "/save-test-screenshot":
                self._send(200, save_test_screenshot(payload))
            elif path == "/task-failed":
                self._send(200, mark_failed(payload))
            elif path == "/reset-failed-tasks":
                self._send(200, reset_failed_tasks())
            elif path == "/reset-running-tasks":
                self._send(200, reset_failed_tasks())
            elif path == "/reset-all-tasks":
                self._send(200, reset_all_tasks())
            else:
                self._send(404, {"ok": False, "error": "not found"})
        except Exception as exc:
            print(f"接口错误 {path}: {exc}")
            self._send(500, {"ok": False, "error": str(exc)})

    def log_message(self, fmt, *args):
        print(f"[{now()}] {self.address_string()} {fmt % args}")


def create_service_server():
    server = ReusableThreadingHTTPServer((HOST, PORT), Handler)
    try:
        init_db()
        created = seed_tasks()
    except Exception:
        server.server_close()
        raise
    print(f"已准备任务，新增 {created} 条")
    print(f"输入 Excel：{INPUT_EXCEL}")
    print(f"结果 Excel：{RESULT_EXCEL}")
    print(f"截图目录：{SCREENSHOT_DIR}")
    print(f"服务地址：http://{HOST}:{PORT}")
    return server


def input_matches_existing_progress(input_path):
    if not DB_PATH.exists() or not RESULT_EXCEL.exists():
        return False

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = read_headers(worksheet)
        question_col = find_column(headers, QUESTION_HEADERS)
        if question_col is None:
            return False
        with connect_db() as conn:
            saved_rows = conn.execute(
                "select row_number, question from tasks group by row_number, question order by row_number"
            ).fetchall()
        if not saved_rows:
            return False
        for saved in saved_rows:
            current = worksheet.cell(row=int(saved["row_number"]), column=question_col).value
            if str(current or "").strip() != str(saved["question"] or "").strip():
                return False
        return True
    finally:
        workbook.close()


def configure_input_excel(input_path):
    global INPUT_EXCEL

    new_path = Path(input_path).expanduser().resolve()
    if not new_path.exists():
        raise FileNotFoundError(f"找不到输入 Excel：{new_path}")

    init_db()
    if input_matches_existing_progress(new_path):
        INPUT_EXCEL = new_path
        return {
            "ok": True,
            "input_excel": str(INPUT_EXCEL),
            "archived_result": "",
            "resumed_existing_progress": True,
            "stats": stats(),
        }

    archive_path = None
    if RESULT_EXCEL.exists():
        archive_dir = OUTPUT_DIR / "archive"
        archive_dir.mkdir(parents=True, exist_ok=True)
        stamp = time.strftime("%Y%m%d_%H%M%S")
        archive_path = archive_dir / f"result_{stamp}.xlsx"
        suffix = 1
        while archive_path.exists():
            archive_path = archive_dir / f"result_{stamp}_{suffix}.xlsx"
            suffix += 1
        shutil.move(str(RESULT_EXCEL), str(archive_path))

    INPUT_EXCEL = new_path
    with lock, connect_db() as conn:
        conn.execute("delete from tasks")

    return {
        "ok": True,
        "input_excel": str(INPUT_EXCEL),
        "archived_result": str(archive_path) if archive_path else "",
        "resumed_existing_progress": False,
    }


def main():
    server = create_service_server()
    print("先启动本服务，再在 Chrome 加载插件并点击开始。")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务已停止")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
